"""
Migrates the seed spreadsheet's free-text 'Terrain Mix' column into three
structured numeric columns (Beginner %, Intermediate %, Advanced %) plus
a data-quality flag, so scoring reads real numbers instead of parsing prose.

Where the numbers came from, per resort:
  - 'sourced': taken directly from a resort's own published piste-colour
    breakdown (the same web research already cited in Data Source).
  - 'sourced_conflicting': published breakdowns exist but disagree
    significantly between sources (see Zermatt) -- a middle estimate is
    used and the conflict is noted explicitly rather than hidden.
  - 'estimated': no numeric breakdown was found; the split is inferred
    from the resort's existing qualitative description (e.g. "mostly
    Beg-Int, limited Adv"). These carry the same honesty as the old
    NEEDS VERIFICATION flag -- treat them as a starting point, not fact.

engine/terrain.py's free-text parser becomes a fallback only, kept for
any future resort added with prose-only terrain data before someone
gets round to sourcing real numbers for it.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SOURCE_PATH = "/home/claude/ski_resort_database_seed.xlsx"
DEST_PATH = "/home/claude/ski_resort_database_seed.xlsx"  # in-place; back up first

# name -> (beginner_pct, intermediate_pct, advanced_pct, quality, note)
TERRAIN_DATA = {
    "Bansko": (40, 40, 20, "sourced", None),
    "Chamonix": (23, 31, 46, "sourced", None),
    "Val Thorens": (29, 61, 10, "sourced", None),
    "St. Anton am Arlberg": (36, 26, 38, "sourced", None),
    "Zermatt": (20, 55, 25, "sourced_conflicting",
                "Published sources disagree substantially -- advanced/black "
                "estimates range from 8% to 35% depending on source. This is "
                "a middle estimate, not a resolved figure."),
    "Livigno": (18, 65, 17, "sourced", None),
    "Kitzbühel": (20, 65, 15, "estimated",
                  "Inferred from a qualitative description, not a published "
                  "numeric breakdown."),
    "Grandvalira (Andorra)": (35, 45, 20, "estimated",
                               "Inferred from a qualitative description, not a "
                               "published numeric breakdown."),
    "Cervinia (Breuil-Cervinia)": (25, 65, 10, "estimated",
                                    "Inferred from a qualitative description, "
                                    "not a published numeric breakdown."),
    "Val Gardena (Selva)": (20, 65, 15, "estimated",
                             "Inferred from a qualitative description, not a "
                             "published numeric breakdown."),
    "Courchevel": (39, 51, 10, "sourced",
                   "Source's own categorization mixes 'blue' and 'Int' labels "
                   "for some runs; treat the beginner/intermediate split as "
                   "approximate even though it's source-derived."),
    "Verbier": (26, 49, 25, "sourced", None),
    "Ischgl": (45, 41, 14, "sourced", None),
    "Sölden": (25, 56, 19, "sourced", None),
    "Serre Chevalier": (30, 40, 30, "estimated",
                         "Inferred from a qualitative description, not a "
                         "published numeric breakdown."),
    "Saalbach-Hinterglemm": (35, 50, 15, "estimated",
                              "Inferred from a qualitative description, not a "
                              "published numeric breakdown."),
    "Alpe d'Huez": (30, 45, 25, "estimated",
                    "Inferred from a qualitative description, not a published "
                    "numeric breakdown."),
    "Pamporovo": (50, 40, 10, "estimated",
                  "Inferred from a qualitative description, not a published "
                  "numeric breakdown."),
    "Poiana Brasov": (40, 45, 15, "estimated",
                       "Inferred from a qualitative description, not a "
                       "published numeric breakdown."),
    "Kranjska Gora": (45, 45, 10, "estimated",
                       "Inferred from a qualitative description, not a "
                       "published numeric breakdown."),
    "Méribel": (58, 30, 12, "sourced", None),
    "Val d'Isère / Tignes": (36, 47, 17, "sourced", None),
    "Obergurgl-Hochgurgl": (30, 55, 15, "estimated",
                             "Inferred from a qualitative description, not a "
                             "published numeric breakdown."),
    "Cortina d'Ampezzo": (35, 50, 15, "estimated",
                           "Inferred from a qualitative description, not a "
                           "published numeric breakdown."),
    "Les Deux Alpes": (65, 23, 12, "sourced", None),
    "Formigal": (30, 55, 15, "estimated",
                 "Inferred from a qualitative description, not a published "
                 "numeric breakdown."),
    "Grand Massif (Flaine)": (53, 37, 10, "sourced", None),
    "Krvavec": (40, 45, 15, "estimated",
                "Inferred from a qualitative description, not a published "
                "numeric breakdown."),
    "Astún-Candanchú": (40, 50, 10, "estimated",
                         "Inferred from a qualitative description, not a "
                         "published numeric breakdown."),
    "Bardonecchia": (35, 45, 20, "estimated",
                      "Inferred from a qualitative description, not a "
                      "published numeric breakdown."),
}

FONT = "Arial"
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")
estimated_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
conflicting_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

OLD_HEADER_ROW = 4

src_wb = openpyxl.load_workbook(SOURCE_PATH, data_only=True)
src_ws = src_wb["SkiResorts"]

# Read all existing rows into memory as plain values first.
old_rows = []
for row in src_ws.iter_rows(min_row=OLD_HEADER_ROW + 1, values_only=True):
    if row[0] is None:
        continue
    old_rows.append(list(row[:20]))

new_wb = openpyxl.Workbook()
ws = new_wb.active
ws.title = "SkiResorts"

new_columns = [
    ("Resort", 16), ("Country", 10), ("Ski Area / Region", 20),
    ("Base Elevation (m)", 10), ("Summit Elevation (m)", 10), ("Vertical Drop (m)", 10),
    ("Number of Lifts", 9), ("Piste Length (km)", 10),
    ("Beginner %", 8), ("Intermediate %", 8), ("Advanced %", 8),
    ("Terrain Data Quality", 12),
    ("Off-Piste Reputation (1-5)", 10), ("Snow Reliability (1-5)", 9),
    ("Nightlife / Apres Rating (1-5)", 10), ("Family Friendliness (1-5)", 9),
    ("Nearest Major Airport(s)", 20), ("Airport-Resort Distance (km)", 10),
    ("Typical Transfer Time", 12), ("Est. Adult 6-Day Ski Pass (EUR)", 11),
    ("Est. Accommodation (EUR/night, mid-range, per room)", 12),
    ("Notes / Skier Profile Fit", 30), ("Data Source(s)", 30),
    ("Terrain Data Note", 30),
]

ws.append(["Ski Resort Database — Phase 1 Seed Data (30 resorts)"])
ws["A1"].font = Font(name=FONT, bold=True, size=14)
ws.append(["Terrain now split into structured Beginner/Intermediate/Advanced % "
           "columns (see Terrain Data Quality + Terrain Data Note for how each "
           "figure was derived). Other ratings (1-5) remain analyst judgment. "
           "Prices are estimates, not live data."])
ws["A2"].font = Font(name=FONT, italic=True, size=9, color="666666")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(new_columns))
ws.append([])

HEADER_ROW = 4
for i, (name, width) in enumerate(new_columns, start=1):
    cell = ws.cell(row=HEADER_ROW, column=i, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_center
    cell.border = border
    ws.column_dimensions[get_column_letter(i)].width = width

for old_row in old_rows:
    (name, country, region, base_e, summit_e, vertical, lifts, piste_km,
     old_terrain_text, off_piste, snow_rel, nightlife, family, airport, dist_km,
     transfer_text, pass_price, accom_price, notes, source) = old_row

    if name not in TERRAIN_DATA:
        raise ValueError(f"No terrain migration data for {name!r} -- add it to TERRAIN_DATA")
    beg, inter, adv, quality, terrain_note = TERRAIN_DATA[name]

    new_row = [
        name, country, region, base_e, summit_e, vertical, lifts, piste_km,
        beg, inter, adv, quality,
        off_piste, snow_rel, nightlife, family,
        airport, dist_km, transfer_text, pass_price, accom_price,
        notes, source, terrain_note or "",
    ]
    ws.append(new_row)

last_row = HEADER_ROW + len(old_rows)
for r in range(HEADER_ROW + 1, last_row + 1):
    quality_val = ws.cell(row=r, column=12).value
    for c in range(1, len(new_columns) + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = Font(name=FONT, size=9)
        cell.alignment = wrap
        cell.border = border
    if quality_val == "estimated":
        for c in range(1, len(new_columns) + 1):
            ws.cell(row=r, column=c).fill = estimated_fill
    elif quality_val == "sourced_conflicting":
        for c in range(1, len(new_columns) + 1):
            ws.cell(row=r, column=c).fill = conflicting_fill

ws.row_dimensions[HEADER_ROW].height = 42
for r in range(HEADER_ROW + 1, last_row + 1):
    ws.row_dimensions[r].height = 60
ws.freeze_panes = f"A{HEADER_ROW + 1}"

# Copy the Legend sheet over, with an added note about the terrain migration.
if "Legend & Notes" in src_wb.sheetnames:
    from copy import copy
    old_legend = src_wb["Legend & Notes"]
    legend = new_wb.create_sheet("Legend & Notes")
    for row in old_legend.iter_rows():
        for cell in row:
            new_cell = legend.cell(row=cell.row, column=cell.column, value=cell.value)
            new_cell.font = copy(cell.font)
            new_cell.alignment = copy(cell.alignment)
    legend.column_dimensions["A"].width = old_legend.column_dimensions["A"].width
    for i in range(1, old_legend.max_row + 1):
        legend.row_dimensions[i].height = old_legend.row_dimensions[i].height

    insert_row = old_legend.max_row + 2
    lines = [
        ("Terrain data migration (this update)", True),
        ("The old free-text 'Terrain Mix' column has been replaced with structured "
         "Beginner %/Intermediate %/Advanced % columns, so scoring reads real numbers "
         "instead of parsing prose. 'Terrain Data Quality' marks each row as 'sourced' "
         "(from a resort's own published breakdown), 'sourced_conflicting' (published "
         "sources disagree -- see Zermatt), or 'estimated' (inferred from a qualitative "
         "description, no numeric source found). 'Terrain Data Note' explains estimated/"
         "conflicting rows specifically. Rows are colour-coded: yellow = estimated, "
         "orange = sourced but conflicting.", False),
    ]
    for i, (text, bold) in enumerate(lines):
        cell = legend.cell(row=insert_row + i, column=1, value=text)
        cell.font = Font(name=FONT, bold=bold, size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        legend.row_dimensions[insert_row + i].height = 45 if not bold else 22

new_wb.save(DEST_PATH)
print(f"Migrated {len(old_rows)} resorts. Saved to {DEST_PATH}")
