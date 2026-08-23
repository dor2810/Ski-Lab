"""
Builds data/transfer_options.xlsx -- the Geneva + Innsbruck research pass.

Follows the design in transfer-subsystem-design.md: one row per
(airport, resort, mode), with cost_basis distinguishing per_person from
per_vehicle. Same data-quality convention as everywhere else in this
project: 'sourced' means a real published rate found this session,
'estimated' means inferred. Nothing is presented as more certain than it is.

Currency note: several operators quote GBP or CHF. Converted at
approximate rates (GBP 1.18, CHF 1.07 to EUR) and flagged in the note --
these are indicative, not locked rates.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")
estimated_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
mandatory_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

COLUMNS = [
    ("Airport", 9), ("Resort", 22), ("Mode", 16),
    ("Cost (EUR)", 10), ("Cost Basis", 12), ("Vehicle Capacity", 9),
    ("Duration (min)", 10), ("Round Trip?", 9), ("Mandatory?", 9),
    ("Runs On Days", 14),
    ("Operator(s)", 20), ("Data Quality", 11), ("Source / Note", 46),
]

# (airport, resort, mode, cost_eur, basis, capacity, duration, round_trip,
#  mandatory, operator, quality, note)
ROWS = [
    # ================= GENEVA (GVA) =================
    ("GVA", "Chamonix", "shared_shuttle", 45, "per_person", None, 75, False, False,
     "Mountain Drop-offs / AlpyBus / Alpine Fleet", "sourced",
     "EUR45pp winter peak, ~EUR35 off-peak; some operators quote from EUR19.50-24.50. "
     "Wide spread -- EUR45 used as a realistic peak-season figure."),
    ("GVA", "Chamonix", "private_transfer", 250, "per_vehicle", 8, 75, False, False,
     "Mountain Drop-offs / Chamonix Valley Transfers", "sourced",
     "EUR200-350 up to 8 depending on operator/vehicle; EUR230-250 each way commonly cited."),
    ("GVA", "Chamonix", "bus", 20, "per_person", None, 90, False, False,
     "FlixBus / coach operators", "sourced",
     "Cheapest option; drops at Chamonix Sud bus station, not door-to-door."),

    ("GVA", "Val Thorens", "shared_shuttle", 50, "per_person", None, 195, False, False,
     "Ben's Bus / Alpine Fleet / AlpNav / Alpskibus", "sourced",
     "One-way from ~EUR42.50-50pp; Ben's Bus GBP94 return (~EUR111). Note Ben's Bus runs "
     "WEEKENDS ONLY on this route -- a real constraint for midweek arrivals.", "SA,SU"),
    ("GVA", "Val Thorens", "private_transfer", 420, "per_vehicle", 8, 180, False, False,
     "Alps2Alps / Ski-Lifts / Cool Runnings", "sourced",
     "Quoted up to ~EUR450 one-way private; EUR350-500 range for sedan/minivan."),

    ("GVA", "Courchevel", "shared_shuttle", 49.50, "per_person", None, 165, False, False,
     "AlpNav / Alpine Fleet", "sourced",
     "EUR44.50-49.50pp depending on operator and fixed-point vs door-to-door."),
    ("GVA", "Courchevel", "private_transfer", 420, "per_vehicle", 8, 150, False, False,
     "Various", "sourced", "EUR350-500 for sedan; minivan adds EUR80-150."),

    ("GVA", "Méribel", "shared_shuttle", 49.50, "per_person", None, 165, False, False,
     "Alpine Fleet / AlpNav", "sourced",
     "From GBP39 (~EUR46); EUR49.50pp typical published rate."),
    ("GVA", "Méribel", "private_transfer", 420, "per_vehicle", 8, 150, False, False,
     "Various", "sourced", "Same Tarentaise corridor pricing as Courchevel."),

    ("GVA", "Verbier", "shared_shuttle", 70, "per_person", None, 155, False, False,
     "Alps2Alps / AlpyBus", "sourced",
     "From EUR70pp (Alps2Alps); ~CHF40-49.50 (~EUR43-53) cited elsewhere. Spread reflects "
     "door-to-door vs fixed-point service."),
    ("GVA", "Verbier", "private_transfer", 320, "per_vehicle", 8, 135, False, False,
     "Various", "sourced",
     "EUR220-350 sedan. NOTE one operator quotes CHF654 (~EUR697) -- premium tier, "
     "not representative. Wide market spread on this route."),
    ("GVA", "Verbier", "train", 60, "per_person", None, 180, False, False,
     "SBB + Le Chable gondola", "estimated",
     "Change at Martigny for Le Chable, then gondola to Verbier. No public bus exists on "
     "this route. Cost inferred from comparable Swiss rail fares."),

    ("GVA", "Zermatt", "train", 110, "per_person", None, 225, False, True,
     "SBB + Matterhorn Gotthard Bahn", "sourced",
     "MANDATORY MODE: Zermatt is CAR-FREE. Road access ends at Tasch; final leg is rail. "
     "SBB to Visp, change to MGB. CHF85-140 second class (~EUR91-150). 3h30-4h total."),
    ("GVA", "Zermatt", "private_transfer", 700, "per_vehicle", 8, 210, False, False,
     "Various", "sourced",
     "EUR600-850 to Tasch only -- CANNOT reach Zermatt itself. Passenger must still take "
     "the shuttle train from Tasch. Expensive and does not remove the rail leg."),

    ("GVA", "Val d'Isère / Tignes", "shared_shuttle", 49.50, "per_person", None, 180, False, False,
     "AlpNav / Alpskibus / Ben's Bus", "sourced",
     "EUR47-49.50pp typical. Alpskibus serves the whole Tarentaise from EUR47."),
    ("GVA", "Val d'Isère / Tignes", "private_transfer", 470, "per_vehicle", 8, 165, False, False,
     "Various", "sourced", "EUR400-550 sedan range cited for this route."),

    ("GVA", "Grand Massif (Flaine)", "shared_shuttle", 40, "per_person", None, 75, False, False,
     "Alpine Fleet / regional operators", "estimated",
     "Not individually sourced this pass. Inferred from comparable ~70km Haute-Savoie "
     "routes (Chamonix/Morzine band). Verify before production use."),
    ("GVA", "Grand Massif (Flaine)", "private_transfer", 240, "per_vehicle", 8, 70, False, False,
     "Various", "estimated",
     "Not individually sourced. Inferred from the Chamonix/Morzine distance band."),

    # ================= INNSBRUCK (INN) =================
    ("INN", "St. Anton am Arlberg", "shared_shuttle", 29.74, "per_person", None, 90, False, False,
     "Alps2Alps / Suntransfers / Ski-Lifts", "sourced",
     "From EUR27pp (Alps2Alps) / EUR29.74 (Suntransfers). 106km, ~75-90min. One of the "
     "best-value transfers in the whole database."),
    ("INN", "St. Anton am Arlberg", "private_transfer", 190, "per_vehicle", 8, 75, False, False,
     "Ski-Lifts / Ted's Transfers / Taxi Alpen", "estimated",
     "Inferred from the Obergurgl private rate (GBP187.87 ~ EUR222) adjusted for the "
     "shorter St Anton run. Not individually sourced."),
    ("INN", "St. Anton am Arlberg", "train", 25, "per_person", None, 105, False, False,
     "OBB (via Innsbruck Hbf)", "estimated",
     "Austria's rail network reaches St Anton directly. Requires airport bus to Innsbruck "
     "Hbf first. Cheap but slower with ski luggage."),

    ("INN", "Ischgl", "shared_shuttle", 29.74, "per_person", None, 105, False, False,
     "Suntransfers / Ski-Lifts / Alps2Alps", "sourced",
     "100km, quoted 75min by Suntransfers but 120min by HolidayTaxis -- conflicting "
     "figures, 105min used as a midpoint."),
    ("INN", "Ischgl", "private_transfer", 200, "per_vehicle", 8, 105, False, False,
     "Holiday Taxi / Taxi Alpen Transfers", "estimated",
     "Max 8 persons per vehicle confirmed by Holiday Taxi; exact rate not published on "
     "the pages reviewed."),

    ("INN", "Sölden", "shared_shuttle", 25.49, "per_person", None, 85, False, False,
     "Suntransfers / Otztal Shuttle", "sourced",
     "80km, 85min, from EUR25.49 (Suntransfers). HolidayTaxis quotes 110min -- "
     "conflicting; the shorter figure matches the stated distance better."),
    ("INN", "Sölden", "private_transfer", 180, "per_vehicle", 8, 85, False, False,
     "Ted's Transfers / Holiday Taxi", "estimated",
     "Inferred from comparable Otztal valley private rates."),

    ("INN", "Obergurgl-Hochgurgl", "shared_shuttle", 39, "per_person", None, 105, False, False,
     "Ski-Lifts / Otztal Shuttle", "sourced",
     "GBP33.06pp (~EUR39) via Ski-Lifts. 97km. Duration quotes vary widely: 87min "
     "(AlpinBus) to 130min (HolidayTaxis) -- 105min used as a midpoint."),
    ("INN", "Obergurgl-Hochgurgl", "private_transfer", 222, "per_vehicle", 8, 100, False, False,
     "Ski-Lifts / AlpinBus / Ted's Transfers", "sourced",
     "GBP187.87 per vehicle (~EUR222) via Ski-Lifts."),
    ("INN", "Obergurgl-Hochgurgl", "bus", 11, "per_person", None, 90, False, False,
     "OBB / regional bus via Innsbruck Hbf", "sourced",
     "EUR10-12 via Solden, departs Innsbruck Hbf. Train to Innsbruck Hbf is EUR18.80 "
     "separately. Cheapest option but weekly schedule (Saturdays) limits usefulness.", "SA"),

    ("INN", "Kitzbühel", "shared_shuttle", 45, "per_person", None, 90, False, False,
     "Ski-Lifts / Monsguide", "estimated",
     "Not individually sourced this pass. Inferred from comparable ~90km Tyrolean routes. "
     "Note SZG (Salzburg) is also viable for Kitzbuhel and may be cheaper."),
    ("INN", "Kitzbühel", "private_transfer", 200, "per_vehicle", 8, 90, False, False,
     "Taxi Alpen Transfers", "estimated", "Not individually sourced."),

    ("INN", "Livigno", "shared_shuttle", 75, "per_person", None, 225, False, False,
     "Various", "estimated",
     "Not individually sourced. Long cross-border run (Austria->Italy) via the Ofenpass "
     "or Reschenpass. BGY (Bergamo) is the more common gateway for Livigno."),
    ("INN", "Livigno", "private_transfer", 500, "per_vehicle", 8, 210, False, False,
     "Various", "estimated", "Not individually sourced. Long-distance cross-border rate."),

    ("INN", "Val Gardena (Selva)", "shared_shuttle", 35.69, "per_person", None, 100, False, False,
     "Suntransfers", "sourced",
     "120km, 100min, from EUR35.69 (Suntransfers, listed as 'Selva di Val Gardena'). "
     "Cross-border Austria->Italy."),
    ("INN", "Val Gardena (Selva)", "private_transfer", 260, "per_vehicle", 8, 100, False, False,
     "Various", "estimated", "Not individually sourced; inferred from the 120km distance band."),

    ("INN", "Cortina d'Ampezzo", "shared_shuttle", 55, "per_person", None, 130, False, False,
     "Various", "estimated",
     "Not individually sourced. VCE (Venice) is the more usual gateway for Cortina; "
     "the INN routing is longer and less served."),
    ("INN", "Cortina d'Ampezzo", "private_transfer", 350, "per_vehicle", 8, 130, False, False,
     "Various", "estimated", "Not individually sourced."),

    # ================= CHAMBERY (CMF) =================
    ("CMF", "Val Thorens", "private_transfer", 300, "per_vehicle", 8, 105, False, False,
     "AlpNav / Alp Venture / 3Valley-transfers", "estimated",
     "CMF is the CLOSEST airport to Val Thorens (~120km, 1h45) but has far fewer flights, "
     "mostly charters. AlpNav explicitly offers PRIVATE ONLY from CMF -- shared availability "
     "is limited here, unlike Geneva. Rate inferred from the shorter distance vs GVA."),
    ("CMF", "Courchevel", "private_transfer", 280, "per_vehicle", 8, 100, False, False,
     "AlpNav / Alp Venture", "estimated",
     "1h40 from Chambery per operator guidance. Private-only pattern applies."),
    ("CMF", "Méribel", "private_transfer", 280, "per_vehicle", 8, 105, False, False,
     "AlpNav / Alp Venture", "estimated", "Same Tarentaise corridor as Courchevel."),
    ("CMF", "Val d'Isère / Tignes", "private_transfer", 320, "per_vehicle", 8, 105, False, False,
     "Alp Venture / 3Valley-transfers", "estimated", "1h45 from Chambery."),

    # ================= GRENOBLE (GNB) =================
    ("GNB", "Alpe d'Huez", "shared_shuttle", 42.50, "per_person", None, 100, False, False,
     "Ben's Bus / Alpine Fleet / AlpNav", "sourced",
     "GBP36 single (~EUR42.50) / GBP63 return via Ben's Bus, ~GBP31 with group discount. "
     "Alpine Fleet EUR35 fixed-point, ~EUR75 door-to-door. GNB is the CLOSEST airport. "
     "Up to 9 buses on Saturdays."),
    ("GNB", "Alpe d'Huez", "private_transfer", 160, "per_vehicle", 7, 100, False, False,
     "Alp Venture / Go Transfer Peak / Abeone-Taxi", "sourced",
     "From EUR160 one-way. 1h40 drive."),
    ("GNB", "Alpe d'Huez", "bus", 20, "per_person", None, 90, False, False,
     "Transisere (public)", "sourced",
     "Public bus several times daily in winter, more at weekends. Not door-to-door; "
     "requires a connection in Grenoble."),
    ("GNB", "Les Deux Alpes", "shared_shuttle", 42.50, "per_person", None, 105, False, False,
     "Ben's Bus / Alpine Fleet", "sourced",
     "GBP36 single / GBP63 return (Ben's Bus). Alpine Fleet EUR35 fixed-point / ~EUR75 "
     "door-to-door, 5 days a week.", "MO,TU,WE,TH,FR"),
    ("GNB", "Les Deux Alpes", "private_transfer", 160, "per_vehicle", 7, 120, False, False,
     "Go Transfer Peak / Alp Venture", "sourced",
     "EUR160 per vehicle up to 7 pax (Go Transfer Peak, WEEKDAYS ONLY special offer). "
     "see2alpes.com also cites ~EUR160.", "MO,TU,WE,TH,FR"),
    ("GNB", "Les Deux Alpes", "bus", 20, "per_person", None, 105, False, False,
     "Transisere (public)", "sourced", "Local bus via Grenoble, ~1h45. Cheapest option."),
    ("GNB", "Serre Chevalier", "shared_shuttle", 50, "per_person", None, 135, False, False,
     "Ben's Bus / Southwest Transfers", "estimated",
     "Ben's Bus lists Serre Chevalier in its network but no rate was quoted on the pages "
     "reviewed. Inferred from the Isere band adjusted for distance."),
    ("GNB", "Serre Chevalier", "private_transfer", 260, "per_vehicle", 8, 135, False, False,
     "Southwest Transfers / Alp Venture", "estimated",
     "Not individually quoted. NOTE Turin (TRN) is also a common gateway for Serre Chevalier."),

    # ================= LYON (LYS) =================
    ("LYS", "Alpe d'Huez", "shared_shuttle", 51, "per_person", None, 135, False, False,
     "Ben's Bus / AlpNav", "sourced",
     "Ben's Bus runs this route on most SATURDAYS ONLY -- a real availability constraint. "
     "GBP43.50 single (~EUR51) / GBP79 return.", "SA"),
    ("LYS", "Alpe d'Huez", "private_transfer", 260, "per_vehicle", 8, 120, False, False,
     "Alp Venture", "estimated", "~2h from Lyon; inferred from the GNB rate plus distance."),
    ("LYS", "Les Deux Alpes", "shared_shuttle", 51, "per_person", None, 135, False, False,
     "Ben's Bus", "sourced",
     "GBP43.50 single / GBP79 return. Runs less regularly than the Grenoble service.", "SA"),
    ("LYS", "Les Deux Alpes", "private_transfer", 260, "per_vehicle", 8, 120, False, False,
     "Alp Venture", "estimated", "~2h from Lyon; rate inferred."),
    ("LYS", "Val d'Isère / Tignes", "shared_shuttle", 55, "per_person", None, 180, False, False,
     "Ben's Bus / AlpNav / Alpskibus", "estimated",
     "Ben's Bus serves Tignes/Val d'Isere from Lyon; exact rate not quoted on the pages "
     "reviewed. Inferred from the Tarentaise band plus distance."),
    ("LYS", "Val d'Isère / Tignes", "private_transfer", 450, "per_vehicle", 8, 165, False, False,
     "Alp Venture / 3Valley-transfers", "estimated", "Long run; rate inferred."),
    ("LYS", "Val Thorens", "shared_shuttle", 42.50, "per_person", None, 180, False, False,
     "Alpine Fleet / AlpyBus / AlpNav", "sourced",
     "Alpine Fleet quotes the SAME starting price from Lyon as from Geneva: GBP36 / "
     "EUR42.50pp. 213km, ~3h. AlpyBus newly running a LYS shared service."),
    ("LYS", "Val Thorens", "private_transfer", 420, "per_vehicle", 8, 180, False, False,
     "Various", "estimated", "Comparable to the Geneva private rate for a similar distance."),
]


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TransferOptions"

    ws.append(["Transfer Options — Geneva (GVA) + Innsbruck (INN) research pass"])
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws.append([
        "One row per (airport, resort, mode). 'Cost Basis' is the key field: per_person "
        "scales with group size, per_vehicle does not — which is what fixes the group-size "
        "flaw in the old distance formula. Yellow = estimated (not individually sourced). "
        "Blue = mandatory mode (no alternative exists). Prices are 2026/27 season "
        "indications, NOT live quotes — re-verify before booking."
    ])
    ws["A2"].font = Font(name=FONT, italic=True, size=9, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    ws.append([])

    header_row = 4
    for i, (name, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=header_row, column=i, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = wrap_center
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = width

    for r in ROWS:
        # r[12] is the optional runs-on-days spec; default "daily".
        runs_on = r[12] if len(r) > 12 else "daily"
        ws.append([
            r[0], r[1], r[2], r[3], r[4], r[5], r[6],
            "Yes" if r[7] else "No", "Yes" if r[8] else "No",
            runs_on,
            r[9], r[10], r[11],
        ])

    last = header_row + len(ROWS)
    for row_i in range(header_row + 1, last + 1):
        quality = ws.cell(row=row_i, column=12).value
        is_mandatory = ws.cell(row=row_i, column=9).value == "Yes"
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_i, column=c)
            cell.font = Font(name=FONT, size=9)
            cell.alignment = wrap
            cell.border = border
            if is_mandatory:
                cell.fill = mandatory_fill
            elif quality == "estimated":
                cell.fill = estimated_fill

    ws.row_dimensions[header_row].height = 40
    for row_i in range(header_row + 1, last + 1):
        ws.row_dimensions[row_i].height = 58
    ws.freeze_panes = f"A{header_row + 1}"

    # --- notes sheet ---
    notes = wb.create_sheet("Notes & Method")
    lines = [
        ("Transfer Options — Notes & Method", True),
        ("", False),
        ("Scope of this pass", True),
        ("Geneva (GVA) and Innsbruck (INN) only — the two highest-leverage airports, covering "
         "16 of the 46 airport-resort pairs in the database (~35%). The remaining 30 pairs across "
         "17 other airports are not yet researched and still fall back to the distance formula.", False),
        ("", False),
        ("Why cost_basis matters", True),
        ("The old model was a single distance formula divided by group_size**0.3, which matches no "
         "real pricing structure. Measured on Val Thorens it predicted EUR412 total for a group of 8. "
         "A private vehicle on that route is roughly EUR420 FLAT — so per person it is ~EUR53, not "
         "EUR51-per-person-times-8. For groups of roughly 4+, per_vehicle modes usually win, and the "
         "old formula could not represent that at all.", False),
        ("", False),
        ("Mandatory modes", True),
        ("Zermatt is car-free. Road access ends at Tasch and the final leg MUST be by rail. A private "
         "transfer to 'Zermatt' actually terminates at Tasch and does not remove the train leg — so "
         "offering it as an alternative to the train would be misleading. This is flagged with "
         "Mandatory=Yes and highlighted blue.", False),
        ("", False),
        ("Conflicting duration data", True),
        ("Several routes have materially conflicting published journey times (Obergurgl: 87min vs "
         "130min; Ischgl: 75min vs 120min; Solden: 85min vs 110min). Midpoints are used and the "
         "conflict is recorded in the note. Do not treat these durations as precise.", False),
        ("", False),
        ("Discrepancies found against the existing resort spreadsheet", True),
        ("Val Thorens: the resort sheet records 150km / 2h30 from Geneva. Transfer operators "
         "consistently state 196km and 3h-3h30. The resort sheet appears to understate this route. "
         "Worth correcting — it feeds the convenience score.", False),
        ("Val Thorens shared shuttles (Ben's Bus) run WEEKENDS ONLY from Geneva. That is a real "
         "availability constraint that no cost model currently captures, and it would silently "
         "mis-price a midweek arrival.", False),
        ("", False),
        ("Currency", True),
        ("Several operators quote GBP or CHF. Converted at approximately GBP 1.18 and CHF 1.07 to "
         "EUR. Indicative only, not locked rates.", False),
        ("", False),
        ("Not yet done", True),
        ("Rental car has not been priced for any pair. It is a per_vehicle mode that may beat private "
         "transfer for longer stays, and is worth adding — but winter Alpine driving (chains, passes, "
         "one-way fees) makes it a genuinely different proposition, not just a cheaper van.", False),
        ("Affiliate/booking URLs are absent. Whether these operators run affiliate programmes has NOT "
         "been verified — it was an assumption, and is flagged as an open question in the design doc "
         "rather than treated as established.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = notes.cell(row=i, column=1, value=text)
        c.font = Font(name=FONT, bold=bold, size=11 if (bold and i == 1) else 10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        notes.row_dimensions[i].height = 42 if not bold else 20
    notes.column_dimensions["A"].width = 115

    out = "/home/claude/ski-trip-optimizer/data/transfer_options.xlsx"
    wb.save(out)
    print(f"Wrote {len(ROWS)} transfer options to {out}")


if __name__ == "__main__":
    build()
