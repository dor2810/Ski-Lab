"""
Rewrites the seed's transfer-time and airport-distance columns from the
MEASURED Google Maps figures in ski_optimizer/data/transfer_drive_times.py.

WHY THIS EXISTS: migration 004 re-pointed six resorts' gateways but
deliberately wrote no timings, because a number for a route nobody had
driven would have been invented. scripts/build_transfer_drive_times.py
has since measured every (resort, gateway) pair in the updated sheet, so
this pass copies those real figures back in.

It also repairs a subtler problem 004 left behind: 'Typical Transfer
Time' still named airports that had just been removed from the airport
column -- Livigno quoted Bergamo, Passo Tonale and Val Gardena quoted
Bolzano. Rebuilding the whole string from the airport column keeps the
two columns from ever disagreeing again.

Runs over EVERY resort, not just the six that changed. A full pass was
how the Méribel error would have been caught: its seed figure claimed
Geneva was 55min / 77.9km away when the resort sits between Courchevel
(128min) and Les Menuires (139min) in the same valley. Verified before
writing -- across the 29 surviving resorts no seed figure now disagrees
with its measurement by more than 10 minutes.

Distance column semantics: it holds ONE number, so it takes the PRIMARY
(left-most) gateway's road distance. When 004 promoted a new primary,
the old value silently began describing a different journey.

Usage: PYTHONPATH=. python3 data/migrations/005_sync_measured_transfers.py
"""
import re
import shutil

import openpyxl

from ski_optimizer.data.transfer_drive_times import DRIVE_TIMES

SEED = "data/ski_resort_database_seed.xlsx"
BACKUP = "data/ski_resort_database_seed.pre_005.xlsx"

COL_RESORT = 1
COL_AIRPORTS = 22
COL_DISTANCE_KM = 23
COL_TRANSFER_TIME = 24

_IATA = re.compile(r"\(([A-Z]{3})\)")


def format_duration(minutes: int) -> str:
    """Match the spreadsheet's existing style: '2h28', or '57min' under an hour."""
    if minutes < 60:
        return f"{minutes}min"
    return f"{minutes // 60}h{minutes % 60:02d}"


def gateway_codes(airport_field: str) -> list:
    """IATA codes in the order listed; left-most is primary."""
    return _IATA.findall(airport_field or "")


def main() -> None:
    shutil.copyfile(SEED, BACKUP)
    print(f"backed up -> {BACKUP}")

    workbook = openpyxl.load_workbook(SEED)
    sheet = workbook["SkiResorts"]

    changed, unchanged, missing = [], 0, []
    for row in range(5, sheet.max_row + 1):
        name = sheet.cell(row, COL_RESORT).value
        if not name:
            continue
        codes = gateway_codes(sheet.cell(row, COL_AIRPORTS).value)

        measurements = []
        for code in codes:
            found = DRIVE_TIMES.get(f"{name}|{code}")
            if found is None:
                missing.append(f"{name}|{code}")
                continue
            measurements.append((code, found))

        if not measurements:
            missing.append(f"{name}: no gateway measured at all")
            continue

        transfer = " / ".join(
            f"{format_duration(found['minutes'])} ({code})" for code, found in measurements
        )
        distance = measurements[0][1]["km"]  # primary gateway

        before = (sheet.cell(row, COL_DISTANCE_KM).value, sheet.cell(row, COL_TRANSFER_TIME).value)
        after = (distance, transfer)
        if before != after:
            sheet.cell(row, COL_DISTANCE_KM).value = distance
            sheet.cell(row, COL_TRANSFER_TIME).value = transfer
            changed.append((name, before, after))
        else:
            unchanged += 1

    if missing:
        raise SystemExit(f"refusing to write, unmeasured routes: {missing}")

    workbook.save(SEED)
    for name, before, after in changed:
        print(f"  {name}")
        print(f"    was: {before[0]}km  {before[1]}")
        print(f"    now: {after[0]}km  {after[1]}")
    print(f"\nupdated {len(changed)} resorts, {unchanged} already correct")


if __name__ == "__main__":
    main()
