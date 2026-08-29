"""
Applies the 2026-08-29 resort/airport verification pass to
data/ski_resort_database_seed.xlsx.

See setup_context_for_claude/RESORT_AIRPORT_REVIEW.md for how each
decision was reached and what evidence backs it.

THREE KINDS OF CHANGE:

1. DROPS (10 resorts). Removed for one of: no bookable transfer at all,
   a duplicate of a resort already in the set, or terrain too small to
   justify flying to it.

2. AIRPORT RE-POINTINGS. Several resorts led with a gateway that has no
   realistic Israeli routing but a flattering transfer time -- Bolzano
   (no direct TLV service) and Bergamo (EUR1698 in a live probe) were
   the main offenders. Airport order in the 'Nearest Major Airport(s)'
   column is meaningful: resort_repository parses it left to right and
   the first entry is treated as primary, so re-ordering is a real
   behavioural change, not cosmetic.

   NOTE ON NUMBERS: this migration only re-orders airports and adds
   gateways. It deliberately writes NO transfer time or distance for a
   newly added gateway. Those are measured by
   scripts/build_transfer_drive_times.py against the Google Maps
   Directions API and written back in a second pass -- inventing a
   plausible-looking duration here is exactly what this project forbids.

3. NOTE CORRECTIONS. Two factual fixes to free-text columns that were
   misleading readers (stale Olympics warning, missing car-free caveat).

Usage: PYTHONPATH=. python3 data/migrations/004_resort_airport_review.py
"""
import shutil

import openpyxl

SEED = "data/ski_resort_database_seed.xlsx"
BACKUP = "data/ski_resort_database_seed.pre_004.xlsx"

COL_RESORT = 1
COL_ISRAELI_ACCESS = 17
COL_AIRPORTS = 22
COL_DISTANCE_KM = 23
COL_TRANSFER_TIME = 24
COL_NOTES = 27

#: Resorts removed by the review, with the reason kept alongside so this
#: file stays readable as a record and not just a list of deletions.
DROPS = {
    "Astún-Candanchú": "3h49 transfer, no Alps2Alps and no Omio coverage",
    "Formigal": "same Pyrenees access problem; no Alps2Alps quote",
    "Vallnord (Pal-Arinsal)": "no transfer coverage; Grandvalira covers Andorra",
    "Poiana Brasov": "no Alps2Alps quote; Bulgaria covered better by Bansko",
    "Krvavec": "30km piste / 521m vertical -- too small to fly to",
    "Sella Ronda (Dolomiti)": "a pass area, not a bookable resort; no transfer coverage",
    "Pamporovo": "real transfer 3h00 via Sofia; Bansko is the better Bulgarian entry",
    "Obergurgl-Hochgurgl": "Sölden dominates it in the same valley and has a quote",
    "Méribel": "near-duplicate of Courchevel; corrupted location data",
    "Avoriaz": "477m vertical (smallest in set), 75km piste, 3/5 snow at EUR140/night",
}

#: resort -> new 'Nearest Major Airport(s)' value. Left-most is primary.
#: Only resorts whose ordering actually changes appear here; Alpe d'Huez,
#: Les Deux Alpes and the Trois Vallées resorts already led with the
#: right gateway and are deliberately untouched.
AIRPORT_REPOINTS = {
    # Toulouse promoted: 1-stop from ~$284 rt and a shorter transfer than
    # Barcelona, which stays as the high-frequency fallback.
    "Grandvalira (Andorra)": "Toulouse (TLS) / Barcelona (BCN)",
    # Munich added so travellers aren't locked into Innsbruck's single
    # Friday Israir rotation.
    "St. Anton am Arlberg": "Innsbruck (INN) / Munich (MUC)",
    # Verona promoted (4 airlines direct); Bolzano dropped -- its 57min
    # transfer was never reachable without a connecting flight.
    "Val Gardena (Selva)": "Verona (VRN) / Innsbruck (INN)",
    # Lyon added as primary: direct Transavia service, where neither
    # Turin nor Grenoble could be confirmed as a workable gateway.
    "Serre Chevalier": "Lyon (LYS) / Turin (TRN) / Grenoble (GNB)",
    # Bergamo dropped (EUR1698 live probe); Malpensa promoted.
    "Livigno": "Milan Malpensa (MXP) / Innsbruck (INN)",
    # Verona promoted over Bergamo; Bolzano dropped.
    "Passo Tonale": "Verona (VRN) / Milan Bergamo (BGY)",
}

#: Gateways added by AIRPORT_REPOINTS that have no measured drive time
#: yet. Listed so the second pass knows what it must fill, and so a
#: reader can see which numbers are deliberately absent.
AWAITING_MEASUREMENT = {
    "St. Anton am Arlberg": "MUC",
    "Serre Chevalier": "LYS",
    "Livigno": "MXP",
}

#: Free-text corrections: (column, new value).
NOTE_FIXES = {
    "Cortina d'Ampezzo": [(
        COL_ISRAELI_ACCESS,
        "Indirect — via Venice (VCE), the cheapest Alpine gateway priced in the "
        "2026-08-29 probe at EUR430 round trip, or Innsbruck. Venice has no direct "
        "TLV service; Innsbruck has one Israir rotation a week on Fridays. The "
        "February 2026 Winter Olympics it co-hosted have now passed: the crowding "
        "and pricing distortion is over, the upgraded infrastructure remains.",
    )],
    "Zermatt": [(
        COL_NOTES,
        "CAR-FREE RESORT — road transfers terminate at Täsch and travellers must "
        "change to the shuttle train for the final leg. Any Alps2Alps or Omio quote "
        "to 'Zermatt' is really a quote to Täsch, so both journey time and cost are "
        "understated until that leg is modelled. Year-round glacier skiing, 5/5 snow "
        "reliability, lift-linked to Cervinia.",
    )],
}


def main() -> None:
    shutil.copyfile(SEED, BACKUP)
    print(f"backed up -> {BACKUP}")

    workbook = openpyxl.load_workbook(SEED)
    sheet = workbook["SkiResorts"]

    # Bottom-up so earlier deletions don't shift the rows still to come.
    removed = []
    for row in range(sheet.max_row, 4, -1):
        name = sheet.cell(row, COL_RESORT).value
        if name in DROPS:
            sheet.delete_rows(row)
            removed.append(name)
    print(f"dropped {len(removed)} resorts")

    missing_drops = set(DROPS) - set(removed)
    if missing_drops:
        raise SystemExit(f"drop targets not found in sheet: {sorted(missing_drops)}")

    repointed, fixed = [], []
    for row in range(5, sheet.max_row + 1):
        name = sheet.cell(row, COL_RESORT).value
        if name in AIRPORT_REPOINTS:
            sheet.cell(row, COL_AIRPORTS).value = AIRPORT_REPOINTS[name]
            # A newly primary gateway invalidates the old primary's
            # distance and transfer string. Blank them rather than carry
            # a number that now describes the wrong journey; the
            # measurement pass refills both.
            if name in AWAITING_MEASUREMENT:
                sheet.cell(row, COL_DISTANCE_KM).value = None
                sheet.cell(row, COL_TRANSFER_TIME).value = None
            repointed.append(name)
        if name in NOTE_FIXES:
            for column, value in NOTE_FIXES[name]:
                sheet.cell(row, column).value = value
            fixed.append(name)

    missing_repoints = set(AIRPORT_REPOINTS) - set(repointed)
    if missing_repoints:
        raise SystemExit(f"repoint targets not found: {sorted(missing_repoints)}")
    missing_fixes = set(NOTE_FIXES) - set(fixed)
    if missing_fixes:
        raise SystemExit(f"note-fix targets not found: {sorted(missing_fixes)}")

    workbook.save(SEED)
    print(f"re-pointed {len(repointed)} airports, corrected {len(fixed)} notes")
    print(f"resorts remaining: {sheet.max_row - 4}")
    print("\nNEXT: PYTHONPATH=. python3 scripts/build_transfer_drive_times.py")
    print("      then write measured distances/times for:", AWAITING_MEASUREMENT)


if __name__ == "__main__":
    main()
