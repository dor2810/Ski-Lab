"""
Adds five new columns to the seed spreadsheet: Avg Annual Snowfall (cm),
Glacier Access, Typical Season, Terrain Park, and Israeli Flight Access.

The last one is deliberately specific to this project's audience -- no
generic ski-resort site publishes "which of these are actually easy to
reach from Tel Aviv," and it's exactly the kind of thing this database
is supposed to know that a generic tool wouldn't.

Same honesty policy as every other migration: 'sourced' means found
this session with a citable source, 'estimated' means inferred from
general knowledge/regional pattern, 'mixed' means some fields in the
bundle are sourced and others estimated (see the Notes column for
which). Nothing here pretends to be more certain than it is.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

SOURCE_PATH = "/home/claude/ski_resort_database_seed.xlsx"
DEST_PATH = "/home/claude/ski_resort_database_seed.xlsx"

# name -> (avg_snowfall_cm, glacier_access, typical_season, terrain_park, israeli_flight_access, quality, note)
EXTENDED_DATA = {
    "Bansko": (280, "No", "~6 Dec\u2013mid Apr", "Yes (small)",
               "Excellent \u2014 Sofia served by low-cost carriers (Wizz Air) plus frequent dedicated "
               "Israeli ski charter packages; the single most popular Israeli ski destination",
               "mixed", "Flight access sourced from Israeli travel sites; snowfall/season estimated."),
    "Chamonix": (420, "Off-piste only (Vall\u00e9e Blanche via Aiguille du Midi); no on-piste glacier lifts",
                 "20 Dec\u20133 May", "Yes",
                 "Indirect \u2014 via Geneva on budget carriers (easyJet); no dedicated Israeli charter",
                 "mixed", "Season dates sourced; snowfall estimated by regional comparison."),
    "Val Thorens": (643, "No \u2014 non-glacial, highest non-glacier base in the Alps (2,300m)",
                    "22 Nov\u20133 May", "Yes",
                    "Indirect \u2014 via Geneva/Chamb\u00e9ry on budget carriers; no dedicated Israeli charter",
                    "sourced", None),
    "St. Anton am Arlberg": (384, "No", "3 Dec\u201319 Apr", "Yes",
                              "Good \u2014 Innsbruck has direct scheduled flights (Israir) this season, "
                              "~1h from resort",
                              "sourced", None),
    "Zermatt": (413, "Yes \u2014 Matterhorn Glacier Paradise, skiable year-round",
                "Year-round (main season late Nov\u2013Apr)", "Yes",
                "Indirect \u2014 via Geneva on budget carriers; no dedicated Israeli charter",
                "sourced", None),
    "Livigno": (300, "No", "~28 Nov\u20134 May", "Yes \u2014 well known for its snow parks",
                "Indirect and long \u2014 via Milan/Bergamo/Innsbruck; no dedicated Israeli charter, "
                "one of the longest transfers in this database",
                "mixed", "Terrain park reputation sourced; snowfall/season estimated."),
    "Kitzb\u00fchel": (300, "No", "12 Dec\u201312 Apr", "Yes",
                    "Good \u2014 Innsbruck/Salzburg both have scheduled direct access",
                    "mixed", "Season dates sourced; snowfall estimated."),
    "Grandvalira (Andorra)": (250, "No", "~early Dec\u2013mid Apr", "Yes",
                               "Indirect only \u2014 via Barcelona/Toulouse with a long onward transfer; "
                               "no dedicated Israeli charter",
                               "estimated", None),
    "Cervinia (Breuil-Cervinia)": (566, "Yes \u2014 linked to Zermatt's glacier", "25 Oct\u20133 May", "Yes",
                                    "Indirect \u2014 via Milan Malpensa; no dedicated Israeli charter",
                                    "mixed", "Snowfall and season dates sourced; terrain park estimated."),
    "Val Gardena (Selva)": (300, "No", "~early Dec\u2013mid Apr", "Yes",
                             "Indirect \u2014 via Innsbruck/Bolzano/Verona; the Innsbruck leg has direct "
                             "scheduled access from Israel",
                             "estimated", None),
    "Courchevel": (418, "No", "5 Dec\u201319 Apr", "Yes",
                   "Indirect \u2014 via Chamb\u00e9ry/Geneva on budget carriers; no dedicated Israeli charter",
                   "sourced", None),
    "Verbier": (469, "Partial \u2014 Mont Fort glacier area, limited glacier terrain",
                "~early Dec\u2013late Apr", "Yes",
                "Indirect \u2014 via Geneva on budget carriers; no dedicated Israeli charter",
                "mixed", "Snowfall sourced; season dates estimated."),
    "Ischgl": (390, "No", "27 Nov\u20133 May", "Yes",
               "Good \u2014 Innsbruck has direct scheduled flights (Israir) this season",
               "sourced_conflicting",
               "Published average-snowfall sources disagree substantially (236cm to 543cm depending on "
               "source) \u2014 this is a rough middle estimate, not a resolved figure. Season dates and "
               "flight access are solidly sourced."),
    "S\u00f6lden": (380, "Yes \u2014 Rettenbach & Tiefenbach glaciers", "2 Oct (glacier)\u20133 May", "Yes",
                "Good \u2014 Innsbruck has direct scheduled flights (Israir) this season",
                "mixed", "Glacier status and season dates sourced; snowfall estimated."),
    "Serre Chevalier": (300, "No", "~early Dec\u2013late Apr", "Yes",
                         "Indirect \u2014 via Turin/Grenoble; no dedicated Israeli charter",
                         "estimated", None),
    "Saalbach-Hinterglemm": (300, "No", "~early Dec\u2013mid Apr", "Yes",
                              "Good \u2014 Salzburg has had seasonal direct flights (Sun D'Or) in recent years",
                              "estimated", None),
    "Alpe d'Huez": (438, "Limited \u2014 small glacier area near Pic Blanc/Sarenne", "6 Dec\u201319 Apr", "Yes",
                    "Indirect \u2014 via Grenoble/Lyon; no dedicated Israeli charter",
                    "sourced", None),
    "Pamporovo": (220, "No", "~mid Dec\u2013early Apr", "Small/unconfirmed",
                  "Good \u2014 Plovdiv/Sofia served by low-cost carriers; Bulgaria is the most "
                  "Israeli-charter-friendly ski region overall",
                  "estimated", None),
    "Poiana Brasov": (200, "No", "~mid Dec\u2013late Mar", "Small/unconfirmed",
                       "Indirect \u2014 via Bucharest; no dedicated Israeli charter, notably less popular "
                       "with Israeli travelers than Bulgaria",
                       "estimated", None),
    "Kranjska Gora": (250, "No", "~mid Dec\u2013late Mar", "Small/unconfirmed",
                       "Indirect \u2014 via Ljubljana; no dedicated Israeli charter",
                       "estimated", None),
    "M\u00e9ribel": (380, "No", "6 Dec\u201317 Apr", "Yes",
                  "Indirect \u2014 via Chamb\u00e9ry/Geneva on budget carriers; no dedicated Israeli charter",
                  "mixed", "Season dates sourced; snowfall estimated by 3 Vall\u00e9es regional comparison."),
    "Val d'Is\u00e8re / Tignes": (625, "Yes \u2014 Grande Motte glacier (Tignes side)",
                              "22 Nov\u20133 May", "Yes",
                              "Indirect \u2014 via Chamb\u00e9ry/Geneva/Lyon; no dedicated Israeli charter",
                              "sourced", "Snowfall is an average of Val d'Is\u00e8re (581cm) and Tignes "
                              "(669cm) 10-year figures \u2014 the two sides differ."),
    "Obergurgl-Hochgurgl": (350, "No \u2014 but very high altitude (3,082m top)",
                             "20 Nov\u2013~20 Apr", "Small/unconfirmed \u2014 quiet, family-oriented resort",
                             "Good \u2014 Innsbruck has direct scheduled flights (Israir) this season",
                             "mixed", "Opening date and flight access sourced; snowfall/closing estimated."),
    "Cortina d'Ampezzo": (350, "No", "29 Nov\u20133 May", "Yes",
                           "Indirect \u2014 via Venice/Innsbruck; co-hosted the 2026 Winter Olympics, "
                           "which may affect pricing/crowds around that period",
                           "mixed", "Season dates sourced; snowfall estimated."),
    "Les Deux Alpes": (309, "Yes \u2014 largest skiable glacier in Europe", "29 Nov\u20133 May", "Yes",
                        "Indirect \u2014 via Grenoble/Lyon; no dedicated Israeli charter",
                        "sourced", None),
    "Formigal": (250, "No", "~early Dec\u2013mid Apr", "Yes",
                 "Indirect and limited \u2014 via Zaragoza or Lourdes (France); the least convenient "
                 "routing from Israel in this database",
                 "estimated", None),
    "Grand Massif (Flaine)": (350, "No", "6 Dec\u201319 Apr", "Yes",
                               "Indirect \u2014 via Geneva on budget carriers; no dedicated Israeli charter",
                               "mixed", "Season dates sourced; snowfall estimated."),
    "Krvavec": (280, "No", "~mid Dec\u2013late Mar", "Small/unconfirmed",
                "Indirect \u2014 via Ljubljana (10km transfer, shortest in this database); no dedicated "
                "Israeli charter",
                "estimated", None),
    "Ast\u00fan-Candanch\u00fa": (250, "No", "~early Dec\u2013mid Apr", "Small/unconfirmed",
                             "Indirect and limited \u2014 via Zaragoza or Pau (France); alongside Formigal, "
                             "the least convenient routing from Israel in this database",
                             "estimated", None),
    "Bardonecchia": (280, "No", "~early Dec\u2013mid Apr", "Yes",
                      "Indirect \u2014 via Turin, the closest major resort to Turin airport in this "
                      "database; no dedicated Israeli charter",
                      "estimated", None),
}

FONT = "Arial"
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")
estimated_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
conflicting_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
mixed_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

OLD_HEADER_ROW = 4

src_wb = openpyxl.load_workbook(SOURCE_PATH, data_only=True)
src_ws = src_wb["SkiResorts"]

old_rows = []
for row in src_ws.iter_rows(min_row=OLD_HEADER_ROW + 1, values_only=True):
    if row[0] is None:
        continue
    old_rows.append(list(row[:24]))  # current layout has 24 columns (through Terrain Data Note)

new_wb = openpyxl.Workbook()
ws = new_wb.active
ws.title = "SkiResorts"

# Insert the 5 new columns right after "Advanced %" / "Terrain Data Quality"
# (positions 9-12 in the old layout), before Off-Piste Reputation.
new_columns = [
    ("Resort", 16), ("Country", 10), ("Ski Area / Region", 20),
    ("Base Elevation (m)", 10), ("Summit Elevation (m)", 10), ("Vertical Drop (m)", 10),
    ("Number of Lifts", 9), ("Piste Length (km)", 10),
    ("Beginner %", 8), ("Intermediate %", 8), ("Advanced %", 8), ("Terrain Data Quality", 12),
    ("Avg Annual Snowfall (cm)", 10), ("Glacier Access", 22), ("Typical Season", 16),
    ("Terrain Park", 16), ("Israeli Flight Access", 34),
    ("Off-Piste Reputation (1-5)", 10), ("Snow Reliability (1-5)", 9),
    ("Nightlife / Apres Rating (1-5)", 10), ("Family Friendliness (1-5)", 9),
    ("Nearest Major Airport(s)", 20), ("Airport-Resort Distance (km)", 10),
    ("Typical Transfer Time", 12), ("Est. Adult 6-Day Ski Pass (EUR)", 11),
    ("Est. Accommodation (EUR/night, mid-range, per room)", 12),
    ("Notes / Skier Profile Fit", 30), ("Data Source(s)", 30),
    ("Terrain Data Note", 30), ("Extended Data Quality", 12), ("Extended Data Note", 30),
]

ws.append(["Ski Resort Database \u2014 Phase 1 Seed Data (30 resorts, extended)"])
ws["A1"].font = Font(name=FONT, bold=True, size=14)
ws.append(["Extended with snowfall, glacier access, season dates, terrain parks, and Israeli flight "
           "access (this update). 'Extended Data Quality' + 'Extended Data Note' explain how each of "
           "the 5 new fields was derived \u2014 see those before treating any of them as fact. Other "
           "ratings/prices remain as documented in earlier notes."])
ws["A2"].font = Font(name=FONT, italic=True, size=9, color="666666")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(new_columns))
ws.append([])

HEADER_ROW = 4
for i, (name, width) in enumerate(new_columns, start=1):
    cell = ws.cell(row=HEADER_ROW, column=i, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_center
    cell.border = border
    ws.column_dimensions[get_column_letter(i)].width = width

for old_row in old_rows:
    (name, country, region, base_e, summit_e, vertical, lifts, piste_km,
     beg_pct, inter_pct, adv_pct, terrain_quality,
     off_piste, snow_rel, nightlife, family, airport, dist_km,
     transfer_text, pass_price, accom_price, notes, source, terrain_note) = old_row

    if name not in EXTENDED_DATA:
        raise ValueError(f"No extended data for {name!r} -- add it to EXTENDED_DATA")
    snowfall, glacier, season, park, flight_access, ext_quality, ext_note = EXTENDED_DATA[name]

    new_row = [
        name, country, region, base_e, summit_e, vertical, lifts, piste_km,
        beg_pct, inter_pct, adv_pct, terrain_quality,
        snowfall, glacier, season, park, flight_access,
        off_piste, snow_rel, nightlife, family,
        airport, dist_km, transfer_text, pass_price, accom_price,
        notes, source, terrain_note, ext_quality, ext_note or "",
    ]
    ws.append(new_row)

last_row = HEADER_ROW + len(old_rows)
EXT_QUALITY_COL = 30  # "Extended Data Quality" column index (1-based)
for r in range(HEADER_ROW + 1, last_row + 1):
    for c in range(1, len(new_columns) + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = Font(name=FONT, size=9)
        cell.alignment = wrap
        cell.border = border
    quality_val = ws.cell(row=r, column=EXT_QUALITY_COL).value
    fill = {"estimated": estimated_fill, "sourced_conflicting": conflicting_fill,
            "mixed": mixed_fill}.get(quality_val)
    if fill:
        for c in range(13, 18):  # only tint the 5 new extended-data columns (13-17)
            ws.cell(row=r, column=c).fill = fill

ws.row_dimensions[HEADER_ROW].height = 46
for r in range(HEADER_ROW + 1, last_row + 1):
    ws.row_dimensions[r].height = 70
ws.freeze_panes = f"A{HEADER_ROW + 1}"

if "Legend & Notes" in src_wb.sheetnames:
    old_legend = src_wb["Legend & Notes"]
    legend = new_wb.create_sheet("Legend & Notes")
    for row in old_legend.iter_rows():
        for cell in row:
            new_cell = legend.cell(row=cell.row, column=cell.column, value=cell.value)
            new_cell.font = copy(cell.font)
            new_cell.alignment = copy(cell.alignment)
    legend.column_dimensions["A"].width = old_legend.column_dimensions["A"].width
    for i in range(1, old_legend.max_row + 1):
        legend.row_dimensions[i].height = old_legend.row_dimensions[i].height

    insert_row = old_legend.max_row + 2
    lines = [
        ("Extended data columns (this update)", True),
        ("Five new columns: Avg Annual Snowfall (cm), Glacier Access, Typical Season, Terrain Park, "
         "and Israeli Flight Access. The last is specific to this project's audience \u2014 no generic "
         "resort site publishes 'is this easy to reach from Tel Aviv,' and it directly reflects the "
         "product's differentiation, not just a nice-to-have data point.", False),
        ("'Extended Data Quality' per row: 'sourced' (snowfall/season genuinely found this session, "
         "cited), 'sourced_conflicting' (published sources disagree \u2014 see Ischgl), 'mixed' (some of "
         "the 5 fields sourced, others estimated \u2014 check 'Extended Data Note' for which), "
         "'estimated' (all 5 fields inferred from regional/altitude pattern, no direct source this "
         "session). Green = mixed, orange = conflicting, yellow = fully estimated, no fill = fully "
         "sourced \u2014 colours apply only to the 5 new columns (13-17), not the whole row.", False),
        ("Glacier Access distinguishes true on-piste/lift-served glacier skiing (Zermatt, S\u00f6lden, "
         "Les Deux Alpes, Tignes' Grande Motte) from off-piste-only glacier access (Chamonix's Vall\u00e9e "
         "Blanche) and partial/limited glacier terrain (Verbier, Alpe d'Huez) \u2014 these are meaningfully "
         "different for a trip-planning purpose and shouldn't be collapsed into one Yes/No.", False),
        ("Israeli Flight Access is a qualitative judgment, not a live route database \u2014 airline "
         "routes change season to season (the Israir Innsbruck route and Sun D'Or Salzburg route "
         "mentioned here are current for the 2026/27 season per Israeli travel sites, not permanent "
         "facts). Treat this column as a starting signal for which resorts are logistically easiest "
         "from Israel, to be re-verified closer to any specific trip's dates \u2014 this is exactly the "
         "kind of live data adapters/flight_adapter.py (Phase 4) should eventually replace.", False),
    ]
    for i, (text, bold) in enumerate(lines):
        cell = legend.cell(row=insert_row + i, column=1, value=text)
        cell.font = Font(name=FONT, bold=bold, size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        legend.row_dimensions[insert_row + i].height = 50 if not bold else 22

new_wb.save(DEST_PATH)
print(f"Extended {len(old_rows)} resorts with 5 new columns. Saved to {DEST_PATH}")
